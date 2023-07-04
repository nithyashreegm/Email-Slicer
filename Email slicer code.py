from tkinter import *
from openpyxl import *
import socket

wb=load_workbook(r"C:\Users\Nithya Shree\Desktop\5th SEM\python project\NewEmailSlice.xlsx")
sheet=wb.active

def excel_file():
    sheet.column_dimensions['A'].width=30
    sheet.column_dimensions['B'].width=30
    sheet.column_dimensions['C'].width=30
    sheet.column_dimensions['D'].width=30

    sheet.cell(row=1,column=1).value='E-mail Name'
    sheet.cell(row=1,column=2).value='User Name'
    sheet.cell(row=1,column=3).value='Domain Name'
    sheet.cell(row=1,column=4).value='IP'

def clear_text():
    text_field.delete(0,END)

def clickme():
    temp_text=text_field.get()
    user_name=temp_text[:temp_text.index('@')]
    domain_name=temp_text[temp_text.index('@')+1:]
    ip_address=socket.gethostbyname(domain_name)
    result=f"Email:{temp_text}\n"\
           f"User Name:{user_name}\n"\
           f"Domain Name:{domain_name}\n"\
           f"IP:{ip_address}\n"
               
    output.insert(END,result)

    if text_field.get()=="":
        print("Empty Line")
    else:
        current_row=sheet.max_row
        current_column=sheet.max_column

        sheet.cell(row=current_row+1,column=1).value=text_field.get()
        sheet.cell(row=current_row+1,column=2).value=user_name
        sheet.cell(row=current_row+1,column=3).value=domain_name
        sheet.cell(row=current_row+1,column=4).value=ip_address

        wb.save(r"C:\Users\Nithya Shree\Desktop\5th SEM\python project\NewEmailSlice.xlsx")
        clear_text()
#driver code
if(__name__ == '__main__'):
    window=Tk()
    window.title("email slicing project")
    window.geometry("800x500")
    excel_file()

    #top level frame
    frame1=Label(window,width=500,height=2,text="email slicing",font=("Bold",20),bg='black',fg='white')
    frame1.pack()
    #top text email writing
    label1=Label(window,text="write an email",font=('Arial',15),fg="#010847")
    label1.pack(pady=15)
    email_name=StringVar()
    #user input box
    text_field=Entry(window,text=email_name,font=('Arial',20),bg='#8dc94d',width='25')
    text_field.pack()
    #slice button
    btn1=Button(window,text="click here to slice",font=("Bold",20),bg='#8591de',command=clickme)
    btn1.pack(pady=40)
    #output box
    output=Text(window,width=50,height=15,bg='#c195e6',fg="black",font=('Arial',15))
    output.pack()
    #exit button
    btn2=Button(window,text="Exit button",font=("Bold",10),bg='red',fg='white',command=window.destroy)
    btn2.pack(pady=20)

window.mainloop()
